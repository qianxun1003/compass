#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
合格实绩分析脚本：从 合格实绩.xlsx 中提取 2022-2024 年合格样本，
按大学·学部·文理分别统计各科目分数段，输出供成绩匹配使用的 JSON 配置。

- 仅使用「合格」样本，过滤「没有出愿」「不考了」「不合格」等。
- 年份权重：2024 > 2023 > 2022（留考难度与报考人数在变）。
- 文理分开建模：文科重日语·文综，理科重数学·理综。
- 输出：各科目 min / p25 / p50 / p75 及样本数 n，便于前端做阈值匹配。
"""

import json
import os
import re
from collections import defaultdict
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("请安装: pip install openpyxl")
    raise

# 项目根目录
ROOT = Path(__file__).resolve().parent.parent
EXCEL_PATH = ROOT / "合格实绩.xlsx"
OUTPUT_JSON = ROOT / "data" / "admission_score_model.json"

# 表名 -> (年份, 文理)。文理: "文" | "理"
SHEET_CONFIG = {
    "2024文科": (2024, "文"),
    "2024理科": (2024, "理"),
    "2023": (2023, None),   # 若 2023 未分表，需从列判断文理
    "2022": (2022, None),
}

# 年份权重（越近年份权重越高）
YEAR_WEIGHT = {2024: 1.0, 2023: 0.8, 2022: 0.6}

# 合格结果的有效取值（表格里可能写「合格」「是」等）
VALID_RESULT = {"合格", "合格 ", "是"}

# 列名可能的别名（与 合格实绩.xlsx 各 sheet 表头对应）
COLUMN_ALIASES = {
    "结果": ["结果", "合否", "是否合格", "最终合格确定", "合格与否", "出愿结果", "status"],
    "大学": ["大学", "学校", "大学名", "报考院校", "school", "name"],
    "学部": ["学部", "学部名", "报考学部", "department"],
    "文理": ["文理", "文/理", "文科理科", "bunri"],
    "日语": ["日语", "日本語", "日语总分", "日语成绩", "EJU日语", "japanese", "jp"],
    "数学1": ["数学1", "数学一", "数学コース1", "文科数学", "数学成绩", "数学", "math1"],
    "数学2": ["数学2", "数学二", "数学コース2", "数学", "math2"],
    "综合": ["综合", "综合科目", "綜合科目", "文综", "文综成绩", "sogo"],
    "物理": ["物理", "physics"],
    "化学": ["化学", "chemistry"],
    "生物": ["生物", "biology"],
    "托福": ["托福", "TOEFL", "英语", "英语分数", "英语成绩", "toefl", "en", "english"],
}


def normalize_header(cell):
    if cell is None:
        return ""
    return str(cell).strip()


def find_column_index(header_row, field_name):
    """根据 COLUMN_ALIASES 找到该字段在 header_row 中的列索引（0-based）。"""
    aliases = COLUMN_ALIASES.get(field_name, [field_name])
    for i, cell in enumerate(header_row):
        h = normalize_header(cell)
        for al in aliases:
            if al in h or h in al:
                return i
    return -1


def parse_number(v):
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v) if v == v else None  # NaN check
    s = str(v).strip().replace(",", "")
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def load_sheet_data(ws, year, default_bunri):
    """从工作表读取所有行，返回 list of dict，仅保留合格样本。"""
    # 不用 read_only 时 ws 才有正确 max_row；逐行读
    max_row = getattr(ws, "max_row", 0) or 0
    if max_row < 2:
        return []
    header_row = [ws.cell(row=1, column=c).value for c in range(1, (ws.max_column or 0) + 1)]
    rows = [header_row]
    for r in range(2, max_row + 1):
        rows.append([ws.cell(row=r, column=c).value for c in range(1, len(header_row) + 1)])
    if len(rows) < 2:
        return []
    result_col = find_column_index(rows[0], "结果")
    if result_col < 0:
        # 没有「结果」列时，可配置为全部视为合格（或跳过）
        pass

    out = []
    for row in rows[1:]:
        if result_col >= 0 and result_col < len(row):
            result_val = normalize_header(str(row[result_col]) if row[result_col] is not None else "")
            # 仅保留合格；排除「没有出愿」「不考了」「不合格」「还没考学」「未知」「否」等
            if result_val:
                skip_keywords = ("不合格", "没有出愿", "不考了", "未出愿", "放弃", "取消", "还没考学", "未知")
                if any(k in result_val for k in skip_keywords):
                    continue
                if result_val == "否":
                    continue
                if "合格" not in result_val and result_val != "是" and result_val != "1":
                    continue  # 2023 表用「是」、2022 表可能用 1 表示合格
        # 合格或无结果列则保留
        record = {"year": year, "bunri": default_bunri}
        for field, _ in COLUMN_ALIASES.items():
            if field == "结果":
                continue
            idx = find_column_index(rows[0], field)
            if idx < 0 or idx >= len(row):
                continue
            val = row[idx]
            if field in ("大学", "学部", "文理"):
                record[field] = normalize_header(val) if val is not None else ""
            elif field in ("日语", "数学1", "数学2", "综合", "物理", "化学", "生物", "托福"):
                num = parse_number(val)
                if num is not None:
                    record[field] = num
        # 文理：若表未分文理，尝试从「文理」列取；否则用 default_bunri
        if default_bunri is None and record.get("文理"):
            b = record["文理"]
            if "文" in b or b == "文科":
                record["bunri"] = "文"
            elif "理" in b or b == "理科":
                record["bunri"] = "理"
            else:
                continue
        else:
            record["bunri"] = record.get("bunri") or default_bunri
        if record.get("大学"):
            out.append(record)
    return out


def weighted_quantile(values_weights, q):
    """values_weights: list of (value, weight). 求加权分位数。"""
    if not values_weights:
        return None
    sorted_vw = sorted(values_weights, key=lambda x: x[0])
    total_w = sum(w for _, w in sorted_vw)
    if total_w <= 0:
        return None
    target = total_w * q
    cum = 0
    for v, w in sorted_vw:
        cum += w
        if cum >= target:
            return v
    return sorted_vw[-1][0]


def stats_for_values(values_weights):
    """返回 min, p25, p50, p75, n（n 为有效样本数）。"""
    if not values_weights:
        return None
    n = len(values_weights)
    w_sum = sum(w for _, w in values_weights)
    if w_sum <= 0:
        return None
    return {
        "min": min(v for v, _ in values_weights),
        "p25": weighted_quantile(values_weights, 0.25),
        "p50": weighted_quantile(values_weights, 0.50),
        "p75": weighted_quantile(values_weights, 0.75),
        "n": n,
    }


# 科目键与 Excel 列名对应
SUBJECT_KEYS = ["日语", "数学1", "数学2", "综合", "物理", "化学", "生物", "托福"]

# 大学名规范化映射：将各种变体统一为标准形式
# 格式：变体名 -> 标准名
UNIVERSITY_NAME_NORMALIZATION = {
    # 繁简统一（统一为简体或日文常用形式）
    "東京大学": "東京大学",
    "东京大学": "東京大学",
    "国際基督教大学": "国際基督教大学",
    "国际基督教大学": "国際基督教大学",
    "国際基督教大学(ICU)": "国際基督教大学",
    "関西学院大学": "関西学院大学",
    "关西学院大学": "関西学院大学",
    "順天堂大学": "順天堂大学",
    "顺天堂大学": "順天堂大学",
    "上智大學": "上智大学",
    "上智大学": "上智大学",
    "二松学舎大学": "二松学舎大学",
    "二松学舍大学": "二松学舎大学",
    "大東文化大学": "大東文化大学",
    "大东文化大学": "大東文化大学",
    "國學院大學": "国学院大学",
    "国学院大学": "国学院大学",
    "東京外国語大学": "東京外国語大学",
    "东京外国语大学": "東京外国語大学",
    "東京外國語大學": "東京外国語大学",
    "関西大学": "関西大学",
    "关西大学": "関西大学",
    "亜細亜大学": "亜細亜大学",
    "亚细亚大学": "亜細亜大学",
    "京都産業大学": "京都産業大学",
    "京都产业大学": "京都産業大学",
    "千葉大学": "千葉大学",
    "千叶大学": "千葉大学",
    "岡山大学": "岡山大学",
    "冈山大学": "岡山大学",
    "岡山商科大学": "岡山商科大学",
    "冈山商科大学": "岡山商科大学",
    "北海道大学": "北海道大学",
    "名古屋大学": "名古屋大学",
    "名古屋市立大学": "名古屋市立大学",
    "名古屋経済大学": "名古屋経済大学",
    "名古屋经济大学": "名古屋経済大学",
    "東京福祉大学": "東京福祉大学",
    "东京福祉大学": "東京福祉大学",
    "東海大学": "東海大学",
    "东海大学": "東海大学",
    "東洋大学": "東洋大学",
    "东洋大学": "東洋大学",
    "聖心女子大学": "聖心女子大学",
    "圣心女子大学": "聖心女子大学",
    "城西国際大学": "城西国際大学",
    "城西国际大学": "城西国際大学",
    "大阪公立大学": "大阪公立大学",
    "大阪産業大学": "大阪産業大学",
    "大阪产业大学": "大阪産業大学",
    "京都先端科学大学": "京都先端科学大学",
    "京都橘大学": "京都橘大学",
    "十文字学園女子大学": "十文字学園女子大学",
    "十文字学园女子大学": "十文字学園女子大学",
    "千葉科学大学": "千葉科学大学",
    "千叶科学大学": "千葉科学大学",
    "国士舘大学": "国士舘大学",
    "国士馆大学": "国士舘大学",
    "埼玉大学": "埼玉大学",
    "琦玉大学": "埼玉大学",
    "一橋大学": "一橋大学",
    "一桥大学": "一橋大学",
    "九州大学": "九州大学",
    "中央大学": "中央大学",
    "中京大学": "中京大学",
    "同志社大学": "同志社大学",
    "南山大学": "南山大学",
    "共立女子大学": "共立女子大学",
    "多摩大学": "多摩大学",
    "嘉悦大学": "嘉悦大学",
    "フェリス女学院大学": "フェリス女学院大学",
    "デジタルハリウッド大学": "デジタルハリウッド大学",
    "北海道教育大学": "北海道教育大学",
    "早稲田大学": "早稲田大学",
    "早稻田大学": "早稲田大学",
    "立命館大学": "立命館大学",
    "立命馆大学": "立命館大学",
    "武蔵野大学": "武蔵野大学",
    "武藏野大学": "武蔵野大学",
    "武蔵野美術大学": "武蔵野美術大学",
    "武藏野美术大学": "武蔵野美術大学",
    "愛知大学": "愛知大学",
    "爱知大学": "愛知大学",
    "福岡女子大学": "福岡女子大学",
    "福冈女子大学": "福岡女子大学",
}


def normalize_university_name(name):
    """
    规范化大学名称：统一繁简、日汉异体、去除括号变体等。
    
    Args:
        name: 原始大学名
        
    Returns:
        规范化后的大学名
    """
    if not name:
        return ""
    
    # 去除首尾空格
    name = name.strip()
    if not name:
        return ""
    
    # 去除常见的括号变体（如 "国際基督教大学(ICU)" -> "国際基督教大学"）
    # 保留括号前的内容
    name = re.sub(r'\([^)]*\)', '', name).strip()
    
    # 去除常见的后缀变体（如 "大学院" -> "大学"）
    name = re.sub(r'大学院$', '大学', name)
    
    # 查找映射表
    normalized = UNIVERSITY_NAME_NORMALIZATION.get(name)
    if normalized:
        return normalized
    
    # 如果没有直接匹配，尝试字符级别的繁简转换（作为后备）
    # 这里可以添加更复杂的繁简转换逻辑，但先使用映射表更可靠
    
    # 返回原名称（如果映射表中没有）
    return name


def normalize_department_name(name):
    """
    规范化学部名称：统一繁简、去除多余空格等。
    
    Args:
        name: 原始学部名
        
    Returns:
        规范化后的学部名
    """
    if not name:
        return ""
    
    # 去除首尾空格
    name = name.strip()
    if not name:
        return ""
    
    # 去除多余的空白字符
    name = re.sub(r'\s+', '', name)
    
    # 可以在这里添加学部名的繁简统一映射，如果需要的话
    # 目前先保持原样，因为学部名变体相对较少
    
    return name


def build_model(all_records):
    """
    按 (大学, 学部, 文理) 分组，对每组各科目做加权统计。
    返回结构：{ "bunka": { "学校名": { "学部名": { "subjects": {...}, "n": int } } }, "rika": {...} }
    """
    # group[(大学, 学部, 文理)] = list of record
    groups = defaultdict(list)
    for r in all_records:
        # 规范化大学名和学部名
        raw_school = r.get("大学", "").strip()
        raw_dept = r.get("学部", "").strip()
        normalized_school = normalize_university_name(raw_school)
        normalized_dept = normalize_department_name(raw_dept)
        
        key = (normalized_school, normalized_dept, r.get("bunri"))
        if not key[0]:
            continue
        groups[key].append(r)

    bunka = {}
    rika = {}

    for (school, dept, bunri), records in groups.items():
        if not school:
            continue
        # 每条记录按年份权重
        weighted = []
        for r in records:
            w = YEAR_WEIGHT.get(r["year"], 0.5)
            weighted.append((r, w))

        subjects = {}
        for sub in SUBJECT_KEYS:
            values_weights = []
            for r, w in weighted:
                v = r.get(sub)
                if v is not None and isinstance(v, (int, float)):
                    values_weights.append((float(v), w))
            st = stats_for_values(values_weights)
            if st and st["n"] > 0:
                subjects[sub] = st

        if not subjects:
            continue

        entry = {
            "subjects": subjects,
            "n": len(records),
        }

        if bunri == "文":
            if school not in bunka:
                bunka[school] = {}
            bunka[school][dept or "(无学部名)"] = entry
        else:
            if school not in rika:
                rika[school] = {}
            rika[school][dept or "(无学部名)"] = entry

    return {"bunka": bunka, "rika": rika}


def main():
    if not EXCEL_PATH.exists():
        print(f"未找到文件: {EXCEL_PATH}")
        print("将生成空模型结构，前端会回退到 recommendJP / recommendEN。")
        model = {"bunka": {}, "rika": {}, "version": "1.0", "generatedAt": "no-data"}
        OUT_DIR = OUTPUT_JSON.parent
        OUT_DIR.mkdir(parents=True, exist_ok=True)
        with open(OUTPUT_JSON, "w", encoding="utf-8") as f:
            json.dump(model, f, ensure_ascii=False, indent=2)
        return

    # 不用 read_only，否则部分 Excel 的 max_row 会错误变成 1
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    all_records = []
    for sheet_name in wb.sheetnames:
        year, default_bunri = SHEET_CONFIG.get(sheet_name, (None, None))
        if year is None:
            m = re.search(r"20(\d{2})", sheet_name)
            year = int(m.group(1)) + 2000 if m else 2023
        ws = wb[sheet_name]
        recs = load_sheet_data(ws, year, default_bunri)
        all_records.extend(recs)
        print(f"  {sheet_name}: 合格样本 {len(recs)} 条")
    wb.close()

    model = build_model(all_records)
    model["version"] = "1.0"
    model["generatedAt"] = __import__("datetime").datetime.now().isoformat()

    OUT_DIR = OUTPUT_JSON.parent
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    with open(OUTPUT_JSON, "w", encoding="utf-8") as f:
        json.dump(model, f, ensure_ascii=False, indent=2)

    print(f"已写入: {OUTPUT_JSON}")
    print(f"  文科 学校数: {len(model['bunka'])}, 理科 学校数: {len(model['rika'])}")


if __name__ == "__main__":
    main()
